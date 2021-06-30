import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font
import string


def automatizar_excel(file_excel):
    data_excel = pd.read_excel(file_excel)
    # data_excel[['Gender','Product line','Total']]
    tp = data_excel.pivot_table(
        index="Gender", columns="Product line", values="Total", aggfunc="sum"
    ).round(0)
    month=file_excel.split('_')[1]
    tp.to_excel(f"sales_{month}", startrow=4, sheet_name="Report")
    wb = load_workbook(f"sales_{month}")
    lb = wb["Report"]
    min_col = wb.active.min_column
    max_col = wb.active.max_column
    min_r = wb.active.min_row
    max_r = wb.active.max_row

    # grafico

    brc = BarChart()
    data = Reference(
        lb, min_col=min_col + 1, max_col=max_col, min_row=min_r, max_row=max_r
    )
    categories = Reference(
        lb, min_col=min_col + 1, max_col=min_col, min_row=min_r + 1, max_row=max_r
    )
    brc.add_data(data, titles_from_data=True)
    brc.set_categories(categories)
    lb.add_chart(brc, "B12")
    brc.title = "Ventas"
    brc.style = 2

    ABC = list(string.ascii_uppercase)
    range_excel = ABC[0:max_col]
    for r in range_excel:
        if r != "A":
            lb[f"{r}{max_col + 1}"] = f"=SUM({r}{min_col+1}:{r}{max_col})"
            lb[f"{r}{max_col+1}"].style = "Currency"
    lb[f"{range_excel[0]}{max_col+1}"] = "TOTAL"
    lb["A1"] = "Reporte"
    m=month.split('.')[0]
    lb["A2"] = f'{m}-2021'
    lb["A1"].font = Font("Arial", bold=True, size=20)
    lb["A2"].font = Font("Arial", bold=True, size=12)
    wb.save(f"sales_{month}")
    return

automatizar_excel('supermarket_junio.xlsx')
